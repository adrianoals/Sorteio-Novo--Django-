from django.shortcuts import render, redirect
from .models import Apartamento, Sorteio, Vaga
import random
from django.utils import timezone
from django.contrib import messages

    
def index(request):
    if request.method == 'POST':
        # Limpar registros anteriores
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos
        apartamentos = list(Apartamento.objects.all())

        # Separar as vagas por bloco
        vagas_a = list(Vaga.objects.filter(vaga__range=('Vaga 27', 'Vaga 50')))
        vagas_b = list(Vaga.objects.filter(vaga__range=('Vaga 01', 'Vaga 26')))
        random.shuffle(vagas_a)  # Embaralhar a lista de vagas para o bloco A
        random.shuffle(vagas_b)  # Embaralhar a lista de vagas para o bloco B

        for apartamento in apartamentos:
            if apartamento.bloco.bloco == 'A' and vagas_a:
                # Atribuir uma vaga do bloco A
                vaga_selecionada = vagas_a.pop()
            elif apartamento.bloco.bloco == 'B' and vagas_b:
                # Atribuir uma vaga do bloco B
                vaga_selecionada = vagas_b.pop()
            else:
                continue  # Se não houver vagas disponíveis, continuar para o próximo apartamento
            Sorteio.objects.create(apartamento=apartamento, vaga=vaga_selecionada)

        # Marcar na sessão que o sorteio foi iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        # Redirecionar para a mesma página para mostrar os resultados do sorteio
        return redirect('index')
    
    else:
        # Verificar se o sorteio já foi iniciado
        sorteio_iniciado = request.session.get('sorteio_iniciado', False)
        resultados_sorteio = []
        vagas_atribuidas = False  # Nova variável para controlar a exibição da mensagem e do botão

        # Carregar a lista de resultados do sorteio apenas se o sorteio já foi iniciado
        if sorteio_iniciado:
            resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').all()
            if resultados_sorteio:
                vagas_atribuidas = True

        return render(request, 'chacara_bom_sucesso/index.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao', '')
        })

# from openpyxl import load_workbook
# from django.http import HttpResponse
# from django.conf import settings
# from .models import Sorteio

# def exportar_para_excel(request):
#     # Construir o caminho completo para o arquivo modelo
#     caminho_modelo = 'static/assets/modelos/sorteio_novo1.xlsx' 

#     # Carregar o workbook do modelo
#     wb = load_workbook(caminho_modelo)
#     ws = wb.active

#     # Obter os resultados do sorteio
#     resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').all()

#     # Supondo que você queira começar a inserir os dados a partir da linha 2
#     linha = 10
#     for sorteio in resultados_sorteio:
#         # Ajuste as colunas conforme a estrutura do seu arquivo modelo
#         ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
#         ws[f'D{linha}'] = sorteio.apartamento.bloco.bloco
#         ws[f'E{linha}'] = sorteio.vaga.vaga
#         linha += 1

#     # Preparar a resposta para enviar o arquivo
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     nome_arquivo = "resultado_sorteio.xlsx"
#     response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

#     # Salvar o workbook modificado no response
#     wb.save(response)

#     return response


def zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('index')
    else:
        return render(request, 'chacara_bom_sucesso/zerar.html')


from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from .models import Sorteio

def exportar_para_excel(request):
    # Construir o caminho completo para o arquivo modelo
    caminho_modelo = 'static/assets/modelos/sorteio_novo1.xlsx' 

    # Carregar o workbook do modelo
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Obter os resultados do sorteio
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').all()

    # Recuperar o horário de conclusão do sorteio da sessão
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')

    # Escrever o horário de conclusão na célula C8
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Supondo que você queira começar a inserir os dados a partir da linha 10
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'D{linha}'] = sorteio.apartamento.bloco.bloco
        ws[f'E{linha}'] = sorteio.vaga.vaga
        linha += 1

    # Preparar a resposta para enviar o arquivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    # Salvar o workbook modificado no response
    wb.save(response)

    return response
