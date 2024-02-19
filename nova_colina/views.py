from django.shortcuts import render

# def nova_colina(request):
#     	return render(request, 'nova_colina/nova_colina.html')


from django.shortcuts import render, redirect
from .models import Apartamento, VagaSimples, VagaDupla, Sorteio
from django.utils import timezone
import random

def nova_colina(request):
    if request.method == 'POST':
        # Limpar registros anteriores
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos
        apartamentos = list(Apartamento.objects.all())

        # Obter todas as vagas simples e duplas disponíveis
        vagas_simples = list(VagaSimples.objects.all())
        vagas_duplas = list(VagaDupla.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas_simples) >= len(apartamentos) and len(vagas_duplas) >= len(apartamentos):
            # Embaralhar as listas de vagas
            random.shuffle(vagas_simples)
            random.shuffle(vagas_duplas)

            # Iterar sobre os apartamentos e atribuir uma vaga simples e uma vaga dupla a cada um
            for apartamento in apartamentos:
                vaga_simples_selecionada = vagas_simples.pop()  # Remover uma vaga simples da lista
                vaga_dupla_selecionada = vagas_duplas.pop()  # Remover uma vaga dupla da lista
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga_simples=vaga_simples_selecionada, 
                    vaga_dupla=vaga_dupla_selecionada
                )
        else:
            # Tratar o caso em que não há vagas suficientes para todos os apartamentos
            # Você pode definir uma mensagem de erro e retorná-la ao template, se desejado
            pass
        
        # Marcar na sessão que o sorteio foi iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        # Redirecionar para a mesma página para mostrar os resultados do sorteio
        return redirect('nova_colina')
    
    else:
        # Verificar se o sorteio já foi iniciado
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga_simples', 'vaga_dupla').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'nova_colina/nova_colina.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao', '')
        })


def zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('nova_colina')
    else:
        return render(request, 'nova_colina/zerar.html')


from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from .models import Sorteio

def excel_nova_colina(request):
    # Construir o caminho completo para o arquivo modelo
    caminho_modelo = 'static/assets/modelos/sorteio_novo2.xlsx' 

    # Carregar o workbook do modelo
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Obter os resultados do sorteio
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga_simples', 'vaga_dupla').all()

    # Recuperar o horário de conclusão do sorteio da sessão
    horario_conclusao_nc = request.session.get('horario_conclusao', 'Horário não disponível')

    # Escrever o horário de conclusão na célula C8
    ws['C8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    # Supondo que você queira começar a inserir os dados a partir da linha 10
    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'C{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'E{linha}'] = sorteio.vaga_simples.vaga_simples
        ws[f'F{linha}'] = sorteio.vaga_dupla.vaga_dupla
        linha += 1

    # Preparar a resposta para enviar o arquivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    # Salvar o workbook modificado no response
    wb.save(response)

    return response
